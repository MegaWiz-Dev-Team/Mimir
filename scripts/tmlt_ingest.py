#!/usr/bin/env python3
"""
Sprint 51 — TMLT (Thai Medical Laboratory Terminology) ingest.

Source: THIS-Center (https://this.or.th/) — `TMLTRF<YYYYMMDD>.zip`.

Smaller cousin of TMT — only 2 concept types + 1 relationship:
  - ITEM:  individual lab tests (~4,758 in v20260501)
  - PANEL: groupings of items (~403 panels)
  - PANELtoITEM: panel-membership pairs (~444)

Uses openpyxl (TMLT ships as .xlsx, not the .xls TMT uses).

Usage:
    python3 scripts/tmlt_ingest.py \\
        --release-dir data/TMLT/TMLTRF20260501 \\
        --source-version tmlt-20260501
"""
from __future__ import annotations
import argparse
import hashlib
import os
import re
import subprocess
import sys
import uuid
from pathlib import Path

from openpyxl import load_workbook


def _have_mysql_cli() -> bool:
    try:
        subprocess.run(["mysql", "--version"], capture_output=True, check=True)
        return True
    except (FileNotFoundError, subprocess.CalledProcessError):
        return False


def mariadb_exec(sql: str) -> str:
    user = os.environ.get("MARIADB_USER", "root")
    pw   = os.environ.get("MARIADB_PASS", "root")
    db   = os.environ.get("MARIADB_DB",   "mimir")
    if _have_mysql_cli():
        host = os.environ.get("MARIADB_HOST", "127.0.0.1")
        port = os.environ.get("MARIADB_PORT", "33306")
        cmd = ["mysql", "-h", host, "-P", port, "-u", user, f"-p{pw}", db, "-B", "-N"]
    else:
        ns = os.environ.get("MARIADB_NAMESPACE", "asgard-infra")
        cmd = ["kubectl", "-n", ns, "exec", "-i", "deploy/mariadb", "--",
               "mariadb", "-u", user, f"-p{pw}", db, "-B", "-N"]
    r = subprocess.run(cmd, input=sql.encode("utf-8"), capture_output=True)
    if r.returncode != 0:
        raise RuntimeError(f"mariadb exec error: {r.stderr.decode()[:500]}")
    return r.stdout.decode("utf-8")


def sql_quote(s: str | None) -> str:
    if s is None or s == "":
        return "NULL"
    return "'" + str(s).replace("\\", "\\\\").replace("'", "\\'") + "'"


# ── Parsers ────────────────────────────────────────────────────────────────


def parse_concept_xlsx(path: Path, concept_type: str) -> list[dict]:
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    iter_rows = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(iter_rows)]
    # Columns: TMLT, FSN, CHANGEDATE
    try:
        id_col = headers.index("TMLT")
    except ValueError:
        id_col = 0
    try:
        fsn_col = headers.index("FSN")
    except ValueError:
        fsn_col = 1
    try:
        chg_col = headers.index("CHANGEDATE")
    except ValueError:
        chg_col = 2

    for row in iter_rows:
        tmlt_id = row[id_col]
        fsn = row[fsn_col]
        chg = row[chg_col] if chg_col < len(row) else None
        if tmlt_id is None or fsn is None:
            continue
        tmlt_id = str(tmlt_id).strip()
        fsn = str(fsn).strip()
        if not tmlt_id or not fsn:
            continue
        change_date = None
        if chg is not None:
            v = str(chg).strip()
            if re.match(r"^\d{8}$", v):
                change_date = f"{v[:4]}-{v[4:6]}-{v[6:]}"
        rows.append({
            "tmlt_id": tmlt_id,
            "concept_type": concept_type,
            "fsn": fsn,
            "change_date": change_date,
        })
    return rows


def parse_relationship_xlsx(path: Path) -> list[tuple[str, str]]:
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    pairs = []
    iter_rows = ws.iter_rows(values_only=True)
    next(iter_rows)  # skip header
    for row in iter_rows:
        if len(row) < 2 or row[0] is None or row[1] is None:
            continue
        a = str(row[0]).strip()
        b = str(row[1]).strip()
        if a and b:
            pairs.append((a, b))
    return pairs


# ── Inserts ────────────────────────────────────────────────────────────────


def insert_codes(rows: list[dict], source_version: str, batch: int = 500) -> int:
    inserted = 0
    for i in range(0, len(rows), batch):
        chunk = rows[i:i + batch]
        values = []
        for r in chunk:
            values.append("({}, {}, {}, {}, {}, NULL, NULL, NOW(), NOW())".format(
                sql_quote(r["tmlt_id"]),
                sql_quote(source_version),
                sql_quote(r["concept_type"]),
                sql_quote(r["fsn"]),
                sql_quote(r["change_date"]),
            ))
        sql = (
            "INSERT INTO tmlt_codes "
            "(tmlt_id, source_version, concept_type, fsn, change_date, "
            " locale_metadata, tenant_id, created_at, updated_at) VALUES "
            + ",".join(values)
            + " ON DUPLICATE KEY UPDATE "
              "concept_type=VALUES(concept_type), fsn=VALUES(fsn), "
              "change_date=VALUES(change_date), updated_at=NOW()"
        )
        mariadb_exec(sql)
        inserted += len(values)
    return inserted


def insert_relationships(pairs: list[tuple[str, str]], source_version: str,
                          batch: int = 500) -> int:
    inserted = 0
    for i in range(0, len(pairs), batch):
        chunk = pairs[i:i + batch]
        values = []
        for panel_id, item_id in chunk:
            values.append("({}, {}, {}, NULL, NOW())".format(
                sql_quote(panel_id), sql_quote(item_id), sql_quote(source_version),
            ))
        sql = (
            "INSERT INTO tmlt_relationships "
            "(panel_id, item_id, source_version, tenant_id, created_at) VALUES "
            + ",".join(values)
            + " ON DUPLICATE KEY UPDATE created_at=created_at"
        )
        mariadb_exec(sql)
        inserted += len(values)
    return inserted


def insert_run(run_id: str, source_version: str, label: str,
               source_url: str | None, sha256: str | None) -> None:
    sql = (
        "INSERT INTO tmlt_ingest_runs "
        "(id, source_version, source_label, source_url, source_sha256, "
        "status, started_at) VALUES "
        f"({sql_quote(run_id)}, {sql_quote(source_version)}, "
        f"{sql_quote(label)}, {sql_quote(source_url)}, "
        f"{sql_quote(sha256)}, 'RUNNING', NOW())"
    )
    mariadb_exec(sql)


def finalize_run(run_id: str, inserted: int, rels: int, status: str, msg: str) -> None:
    sql = (
        "UPDATE tmlt_ingest_runs SET "
        f"rows_inserted={inserted}, rows_relationships={rels}, "
        f"status={sql_quote(status)}, status_message={sql_quote(msg[:1000])}, "
        "finished_at=NOW() "
        f"WHERE id={sql_quote(run_id)}"
    )
    mariadb_exec(sql)


# ── Main ───────────────────────────────────────────────────────────────────


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--release-dir", required=True, type=Path,
                    help="Path to TMLTRF<YYYYMMDD>/ root")
    ap.add_argument("--source-version", required=True,
                    help="e.g. tmlt-20260501")
    ap.add_argument("--source-url", default="https://this.or.th/")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not args.release_dir.exists():
        print(f"ERR: {args.release_dir} not found", file=sys.stderr)
        return 1

    # Find Concept + Relationship dirs (release wraps files in a sub-folder)
    concept_dir = None
    rel_dir = None
    for p in args.release_dir.rglob("*_BONUS/Concept"):
        if p.is_dir(): concept_dir = p; break
    for p in args.release_dir.rglob("*_BONUS/Relationship"):
        if p.is_dir(): rel_dir = p; break
    if not concept_dir or not rel_dir:
        print(f"ERR: Concept/Relationship dirs not found under {args.release_dir}",
              file=sys.stderr)
        return 1

    print(f"=== TMLT ingest from {args.release_dir} ===")
    print(f"  Concept dir:      {concept_dir}")
    print(f"  Relationship dir: {rel_dir}")

    # 2 concept files: TMLT_ITEM*.xlsx and TMLT_PANEL*.xlsx
    all_rows = []
    for ct, prefix in [("ITEM", "TMLT_ITEM"), ("PANEL", "TMLT_PANEL")]:
        matches = list(concept_dir.glob(f"{prefix}*.xlsx"))
        if not matches:
            print(f"  WARN: no file for {prefix}", file=sys.stderr); continue
        rows = parse_concept_xlsx(matches[0], ct)
        print(f"  {ct:5s}: {len(rows):>6,} rows from {matches[0].name}")
        all_rows.extend(rows)
    print(f"  TOTAL concepts: {len(all_rows):,}")

    # 1 relationship file: PANELtoITEM*.xlsx
    pairs: list[tuple[str, str]] = []
    rel_files = list(rel_dir.glob("PANELtoITEM*.xlsx"))
    if rel_files:
        pairs = parse_relationship_xlsx(rel_files[0])
        print(f"  PANEL→ITEM: {len(pairs):,} pairs from {rel_files[0].name}")
    else:
        print(f"  WARN: PANELtoITEM file not found", file=sys.stderr)

    if args.dry_run:
        print("\n[dry-run] skipping DB writes")
        return 0

    # Hash all the xlsx
    sha = hashlib.sha256()
    for fn in sorted(concept_dir.glob("*.xlsx")) + sorted(rel_dir.glob("*.xlsx")):
        sha.update(fn.read_bytes())
    sha_hex = sha.hexdigest()

    run_id = str(uuid.uuid4())
    print(f"\n=== Audit run: {run_id} ===")
    insert_run(run_id, args.source_version, args.source_version,
               args.source_url, sha_hex)

    try:
        print(f"\n=== Inserting {len(all_rows):,} concept rows ===")
        inserted = insert_codes(all_rows, args.source_version)
        print(f"  ✓ inserted/upserted: {inserted:,}")

        rel_count = 0
        if pairs:
            print(f"\n=== Inserting {len(pairs):,} relationships ===")
            rel_count = insert_relationships(pairs, args.source_version)
            print(f"  ✓ relationships: {rel_count:,}")

        finalize_run(run_id, inserted, rel_count, "COMPLETED",
                     f"TMLT release {args.source_version} ingested")
    except Exception as e:
        finalize_run(run_id, 0, 0, "FAILED", str(e)[:1000])
        raise

    # Verify
    print("\n=== Verify ===")
    print(mariadb_exec(
        f"SELECT concept_type, COUNT(*) FROM tmlt_codes "
        f"WHERE source_version='{args.source_version}' GROUP BY concept_type"
    ))
    print(mariadb_exec(
        f"SELECT COUNT(*) AS panel_item_links FROM tmlt_relationships "
        f"WHERE source_version='{args.source_version}'"
    ))
    return 0


if __name__ == "__main__":
    sys.exit(main())
